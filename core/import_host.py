from openpyxl import load_workbook
import os


class ImportHost:
    def __init__(self, file_name, category):
        self.file_name = file_name
        self.category = category

    def open_file(self):
        app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', ''))
        upload_folder = os.path.join(app_root, 'data')
        wb = load_workbook(filename=os.path.join(upload_folder, self.file_name), read_only=True)
        try:
            self.ws = wb[self.category]
            return self.ws
        except KeyError:
            if self.category == "":
                self.errors = "You have to enter sheet name."
                return 1
            else:
                self.errors = "Sheet \"" + self.category + "\" not found."
                return 1

    def read_file(self):
        max_col = self.ws.max_column
        max_row = self.ws.max_row
        file_data = []
        self.file_data = file_data
        for i in range(1, max_col + 1):
            temp = []
            for j in range(1, max_row + 1):
                value = self.ws.cell(row=j, column=i).value
                temp.append(value)
            self.file_data.append(temp)

    def print_file(self):
        return self.file_data

    def print_errors(self):
        return self.errors

    def is_valid(self):
        max_col = self.ws.max_column
        first_row = []
        pattern = ['name', 'description', 'address', 'snmp_version', 'community', 'security_name',
                   'security_level', 'auth_protocol',  'auth_key', 'priv_protocol', 'priv_key', 'interface', 'uptime',
                   'chassis_temperature', 'fan_status', 'cpu_utilization', 'is_on']
        for i in range(max_col):
            first_row.append(self.file_data[i][0])
        if first_row != pattern:
            return 1

    def is_empty(self):
        errors = []
        v_2c = []
        v_3 = []

        for i in [0, 2, 3]:
            if None in self.file_data[i]:
                row_number = self.file_data[i].index(None) + 1
                col_name = self.file_data[i][0]
                errors.append(col_name + " is empty in " + str(row_number) + ". row.")

        for i in range(self.ws.max_row):
            if "2c" in [self.file_data[3][i]]:
                v_2c.append(i)
            elif 3 in [self.file_data[3][i]]:
                v_3.append(i)

        for i in v_2c:
            if None in [self.file_data[4][i]]:
                col_name2 = self.file_data[4][0]
                row_number2 = i + 1
                errors.append(col_name2 + " is empty in " + str(row_number2) + ". row.")

        for i in v_3:
            for j in range(5, 11):
                if None in [self.file_data[j][i]]:
                    col_name3 = self.file_data[j][0]
                    row_number3 = i + 1
                    errors.append(col_name3 + " is empty in " + str(row_number3) + ". row.")
        if errors:
            return 1
