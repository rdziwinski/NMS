from openpyxl import load_workbook
import os
from flask import Flask
import json


class Xlsx():
    def __init__(self, file_name):

        self.file_name = file_name

        wb = load_workbook(filename=os.path.dirname(os.path.abspath(__file__)) + "/" + file_name, read_only=True)
        ws = wb.worksheets[0]
        self.max_col = ws.max_column
        self.max_row = ws.max_row
        self.database = ws
        self.result = []
        if self.max_col == 11:
            for i in range(1, self.max_col+1):
                temp = []
                for j in range(1, self.max_row+1):
                    value = self.database.cell(row=j, column=i).value
                    temp.append(value)
                self.result.append(temp)

    def get_data(self):
        return self.result

    def is_ok(self):
        if self.result != []:
            first_row = []
            pattern = ['name', 'description', 'address', 'snmp_version', 'community', 'security_name',
                       'security_level', 'auth_protocol', 'priv_key', 'priv_protocol', 'auth_key']
            for i in range(self.max_col):
                first_row.append(self.result[i][0])
            if first_row != pattern:
                return 0
        else:
            return 0

    def is_empty(self):
        is_empty = []
        v_2c = []
        v_3 = []

        for i in [0, 2, 3]:
            if None in self.result[i]:
                row_number = self.result[i].index(None) + 1
                col_name = self.result[i][0]
                is_empty.append(col_name + " is empty in " + str(row_number) + ". row.")

        for i in range(self.max_row):
            if "2c" in [self.result[3][i]]:
                v_2c.append(i)
            elif 3 in [self.result[3][i]]:
                v_3.append(i)

        for i in v_2c:
            if None in [self.result[4][i]]:
                col_name2 = self.result[4][0]
                row_number2 = i + 1
                is_empty.append(col_name2 + " is empty in " + str(row_number2) + ". row.")

        for i in v_3:
            for j in range(5, 11):
                if None in [self.result[j][i]]:
                    col_name3 = self.result[j][0]
                    row_number3 = i + 1
                    is_empty.append(col_name3 + " is empty in " + str(row_number3) + ". row.")
        return is_empty


# Routers = Xlsx("hosts.xlsx")
# result = Routers.is_empty()
# print(result)

# print(temp)
# if result == []:
#     print("Wszystko jest")
# else:
#     for i in result:
#         print(i)


# for i in range(max_row-1):
#     host = Host(name=data[0][i], description=data[1][i], address=data[2][i], snmp_version=data[3][i], community=data[4][i],
#          security_name=data[5][i], security_level=data[6][i], auth_protocol=data[7][i], priv_key=data[8][i],
#          priv_protocol=data[9][i], auth_key=data[10][i]).add()
