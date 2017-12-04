from openpyxl import load_workbook
import os


class ImportFile:
    def __init__(self, file_name, category):
        self.file_name = file_name
        self.category = category

    def open_file(self):
        app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', ''))
        upload_folder = os.path.join(app_root, 'data')
        wb = load_workbook(filename=os.path.join(upload_folder, self.file_name), read_only=True)
        try:
            ws = wb[self.category]
            self.ws = ws
            return self.ws
        except KeyError:
            if self.category == "":
                self.errors = "You have to enter sheet name."
                return 1
            else:
                self.errors = "Sheet \"" + self.category + "\" not found."
                return 1

    def read_file(self):
        max_col = self.ws.max_column
        max_row = self.ws.max_row
        file_data = []
        self.file_data = file_data
        for i in range(1, max_col + 1):
            temp = []
            for j in range(1, max_row + 1):
                value = self.ws.cell(row=j, column=i).value
                temp.append(value)
            self.file_data.append(temp)

    def print_file(self):
        return self.file_data

    def print_errors(self):
        return self.errors
